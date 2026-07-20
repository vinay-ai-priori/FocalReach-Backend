"""Lead file (CSV/XLSX) parsing, validation, and confirmed import into companies + leads.

v2 schema: ownership is a single path (lead_import → campaign → user → org). Companies
are canonical per (organization, domain/name) and upserted — never duplicated per run;
per-run verdicts live in CompanyQualification. Raw rows are staged in lead_import_rows
and purged on confirm."""

import csv
import io
import re

from sqlalchemy import insert, select
from sqlalchemy.orm import Session

from app.core.exceptions import ValidationFailedError
from app.core.logging import get_logger
from app.models.campaign import Campaign
from app.models.company import Company, CompanyQualification
from app.models.lead import Lead
from app.models.lead_import import ImportKind, ImportStatus, LeadImport, LeadImportRow
from app.repositories.lead_import_repository import LeadImportRepository
from app.services.csv.column_matcher import build_missing_field_report, match_columns
from app.services.csv.dedup_service import build_dedup_index
from app.services.website.url_validator import extract_domain

logger = get_logger(__name__)

MAX_ROWS = 20000


def _parse_csv(file_bytes: bytes) -> tuple[list[str], list[dict]]:
    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_bytes.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValidationFailedError("The CSV file has no header row.")

    columns = [c.strip() for c in reader.fieldnames if c and c.strip()]
    rows = []
    for i, row in enumerate(reader):
        if i >= MAX_ROWS:
            raise ValidationFailedError(f"The file exceeds the {MAX_ROWS} row limit for a single import.")
        rows.append({(k or "").strip(): (v or "").strip() for k, v in row.items()})
    return columns, rows


def _parse_xlsx(file_bytes: bytes) -> tuple[list[str], list[dict]]:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        raise ValidationFailedError("The Excel file could not be read — it may be corrupted or password-protected.")
    try:
        sheet = workbook.active
        if sheet is None:
            raise ValidationFailedError("The Excel file has no worksheets.")
        rows_iter = sheet.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header or not any(c is not None and str(c).strip() for c in header):
            raise ValidationFailedError("The Excel file has no header row.")
        columns = [str(c).strip() if c is not None else "" for c in header]

        rows: list[dict] = []
        for i, values in enumerate(rows_iter):
            if i >= MAX_ROWS:
                raise ValidationFailedError(f"The file exceeds the {MAX_ROWS} row limit for a single import.")
            if values is None or all(v is None or not str(v).strip() for v in values):
                continue  # skip fully blank rows (common trailing artifact in spreadsheets)
            rows.append({
                col: (str(values[idx]).strip() if idx < len(values) and values[idx] is not None else "")
                for idx, col in enumerate(columns)
                if col
            })
        columns = [c for c in columns if c]
        return columns, rows
    finally:
        workbook.close()


def organization_id_of(lead_import: LeadImport) -> int | None:
    """The single ownership path: import → campaign → user → organization.
    None = the super admin's own space (they sit outside every organization)."""
    return lead_import.campaign.user.organization_id


def raw_rows_of(db: Session, lead_import: LeadImport) -> list[dict]:
    """Staged raw rows, in file order. Empty once the import is confirmed (purged)."""
    return list(
        db.scalars(
            select(LeadImportRow.data)
            .where(LeadImportRow.lead_import_id == lead_import.id)
            .order_by(LeadImportRow.row_number)
        )
    )


def parse_and_validate(
    db: Session,
    campaign: Campaign,
    filename: str,
    file_bytes: bytes,
    kind: ImportKind = ImportKind.PRIMARY,
) -> LeadImport:
    if (filename or "").lower().endswith(".xlsx"):
        columns, rows = _parse_xlsx(file_bytes)
    else:
        columns, rows = _parse_csv(file_bytes)
    columns = [c for c in columns if c]
    # The raw bytes are no longer needed once parsed into `rows` — release them now so
    # the file isn't held in RAM twice for the rest of the request (matching, DB write).
    del file_bytes

    if not rows:
        raise ValidationFailedError("The file contains no data rows.")

    matching = match_columns(columns, db=db)
    column_mapping = {key: val["csv_column"] for key, val in matching.items()}
    mapping_meta = {
        key: {"confidence": val["confidence"], "source": val.get("source")}
        for key, val in matching.items()
        if val["csv_column"]
    }
    missing = build_missing_field_report(matching)

    lead_import = LeadImport(
        campaign_id=campaign.id,
        kind=kind,
        filename=filename,
        status=ImportStatus.MAPPING,
        total_rows=len(rows),
        raw_columns=columns,
        column_mapping=column_mapping,
        mapping_meta=mapping_meta,
        missing_fields=missing,
    )
    lead_import = LeadImportRepository(db).create(lead_import)
    # Core bulk insert (insertmanyvalues): batches thousands of rows into a handful of
    # round trips — the ORM add_all path costs one round trip per row against a remote DB.
    db.execute(
        insert(LeadImportRow),
        [{"lead_import_id": lead_import.id, "row_number": i, "data": row} for i, row in enumerate(rows)],
    )
    db.commit()
    return lead_import


def compute_stats(db: Session, lead_import: LeadImport, rows: list[dict] | None = None) -> dict:
    """CSV analytics for the validation page, computed from the current column mapping so
    they refresh whenever the user remaps a column. Pass `rows` when the caller already
    fetched them — the staged payload is large and re-fetching it is a full round trip."""
    if rows is None:
        rows = raw_rows_of(db, lead_import)
    mapping = lead_import.column_mapping or {}

    companies: set[str] = set()
    kept = 0
    drop_no_company = drop_no_name = drop_no_email = 0

    for row in rows:
        verdict = classify_row(row, mapping)
        if verdict == "no_company":
            drop_no_company += 1
            continue
        if verdict == "no_name":
            drop_no_name += 1
            continue
        if verdict == "no_email":
            drop_no_email += 1
            continue
        # kept row
        kept += 1
        companies.add(_get(row, mapping, "company_name").lower())

    return {
        "rows_detected": lead_import.total_rows,
        "columns_detected": len(lead_import.raw_columns or []),
        "unique_companies": len(companies),
        "total_leads": kept,  # rows that will actually be imported
        "rows_dropped": drop_no_company + drop_no_name + drop_no_email,
        "dropped_missing_company": drop_no_company,
        "dropped_missing_name": drop_no_name,
        "dropped_missing_email": drop_no_email,
    }


def update_mapping(db: Session, lead_import: LeadImport, column_mapping: dict[str, str | None]) -> LeadImport:
    valid_columns = set(lead_import.raw_columns)
    for key, column in column_mapping.items():
        if column is not None and column not in valid_columns:
            raise ValidationFailedError(f"Column '{column}' does not exist in the uploaded CSV.")
    merged = {**lead_import.column_mapping, **column_mapping}
    meta = dict(lead_import.mapping_meta or {})
    for key, column in column_mapping.items():
        if column:
            meta[key] = {"confidence": 100.0, "source": "manual"}
        else:
            meta.pop(key, None)
    matching = {k: {"csv_column": v, "confidence": 100.0 if v else 0.0} for k, v in merged.items()}
    missing = build_missing_field_report(matching)
    return LeadImportRepository(db).update(
        lead_import, column_mapping=merged, mapping_meta=meta, missing_fields=missing
    )


def _get(row: dict, mapping: dict, key: str) -> str | None:
    column = mapping.get(key)
    if not column:
        return None
    value = (row.get(column) or "").strip()
    return value or None


def _parse_int(value: str | None) -> int | None:
    if not value:
        return None
    digits = re.sub(r"[^\d]", "", value)
    return int(digits) if digits else None


def classify_row(row: dict, mapping: dict) -> str:
    """A row is only imported when it can be identified and contacted.
    Returns 'keep' or a drop reason: 'no_company' | 'no_name' | 'no_email'."""
    if not _get(row, mapping, "company_name"):
        return "no_company"
    if not _get(row, mapping, "full_name"):
        return "no_name"
    if not _get(row, mapping, "email"):
        return "no_email"
    return "keep"


def _company_key(name: str, domain: str | None) -> str:
    """Canonical identity within an org: domain when present, else normalized name."""
    return domain.lower() if domain else f"name:{name.lower()}"


def _upsert_company(
    db: Session,
    organization_id: int | None,
    row: dict,
    mapping: dict,
    company_name: str,
    website: str | None,
    domain: str | None,
) -> Company:
    """Find-or-create the canonical company for this org; refresh firmographics from
    the CSV on hit (newer upload wins for CSV-sourced fields, enrichment untouched)."""
    stmt = select(Company).where(
        Company.organization_id.is_(None) if organization_id is None else Company.organization_id == organization_id
    )
    stmt = stmt.where(Company.domain == domain) if domain else stmt.where(
        Company.domain.is_(None), Company.name.ilike(company_name)
    )
    company = db.scalars(stmt).first()
    firmographics = dict(
        name=company_name,
        website=website,
        domain=domain,
        industry=_get(row, mapping, "company_industry"),
        description=_get(row, mapping, "company_description"),
        city=_get(row, mapping, "company_city"),
        state=_get(row, mapping, "company_state"),
        country=_get(row, mapping, "company_country"),
        employee_count=_parse_int(_get(row, mapping, "company_employee_count")),
        employee_range=_get(row, mapping, "company_employee_range"),
        annual_revenue=_get(row, mapping, "company_revenue"),
        linkedin_url=_get(row, mapping, "company_linkedin"),
    )
    if company is None:
        company = Company(organization_id=organization_id, **firmographics)
        db.add(company)
        db.flush()
    else:
        for field, value in firmographics.items():
            if value is not None:
                setattr(company, field, value)
    return company


def _materialize_rows(
    db: Session,
    target_import: LeadImport,
    rows: list[dict],
    mapping: dict,
    dedup_index,
    companies_by_key: dict[str, Company],
    skip_identity=None,  # CampaignIdentityIndex | None — skip rows already in the campaign
) -> tuple[int, int, int]:
    """Turn raw rows into canonical Company + CompanyQualification + Lead records on
    `target_import`. Returns (created_leads, dropped, cross_campaign_duplicates)."""
    organization_id = organization_id_of(target_import)
    leads: list[Lead] = []
    dropped = 0
    duplicates = 0
    seen_emails: set[str] = set()
    qualified_company_ids: set[int] = {
        q.company_id for q in target_import.company_qualifications
    }

    for row in rows:
        # Drop rows we can't identify or contact (missing company name, name, or email).
        if classify_row(row, mapping) != "keep":
            dropped += 1
            continue

        company_name = _get(row, mapping, "company_name")
        website = _get(row, mapping, "company_website")
        domain = None
        if website:
            try:
                domain = extract_domain(website)
            except Exception:
                domain = None
        full_name = _get(row, mapping, "full_name")
        email = _get(row, mapping, "email")

        # In-file duplicate contact (DB also enforces via ux_leads_import_email).
        if email.lower() in seen_emails:
            dropped += 1
            continue

        # Row-level identity: skip contacts already present in this campaign (re-upload).
        if skip_identity is not None and skip_identity.lead_exists(company_name, domain, email, full_name):
            continue
        seen_emails.add(email.lower())

        key = _company_key(company_name, domain)
        company = companies_by_key.get(key)
        if company is None:
            company = _upsert_company(db, organization_id, row, mapping, company_name, website, domain)
            companies_by_key[key] = company
        if company.id not in qualified_company_ids:
            db.add(CompanyQualification(lead_import_id=target_import.id, company_id=company.id))
            qualified_company_ids.add(company.id)

        # We only collect Full Name; derive first/last from it for greetings & personalization.
        name_parts = full_name.split()
        first_name = name_parts[0] if name_parts else None
        last_name = " ".join(name_parts[1:]) if len(name_parts) > 1 else None

        is_duplicate, dup_reason, _ = dedup_index.evaluate(company.name, company.domain, email, full_name)
        if is_duplicate:
            duplicates += 1

        leads.append(
            Lead(
                lead_import_id=target_import.id,
                company_id=company.id,
                full_name=full_name,
                first_name=first_name,
                last_name=last_name,
                title=_get(row, mapping, "title"),
                seniority=_get(row, mapping, "seniority"),
                department=_get(row, mapping, "department"),
                email=email,
                phone=_get(row, mapping, "phone"),
                linkedin_url=_get(row, mapping, "linkedin_url"),
                country=_get(row, mapping, "contact_country"),
                time_in_role=_get(row, mapping, "time_in_role"),
                time_at_company=_get(row, mapping, "time_at_company"),
                years_experience=_get(row, mapping, "years_experience"),
                is_duplicate=is_duplicate,
                duplicate_reason=dup_reason,
            )
        )

    db.add_all(leads)
    return len(leads), dropped, duplicates


def _purge_raw_rows(db: Session, lead_import: LeadImport) -> None:
    db.query(LeadImportRow).filter(LeadImportRow.lead_import_id == lead_import.id).delete(
        synchronize_session=False
    )


def confirm_import(db: Session, lead_import: LeadImport) -> LeadImport:
    """Materialize staged rows into canonical Company + Lead records."""
    from app.services.csv.reupload_service import icp_fingerprint

    if lead_import.status != ImportStatus.MAPPING:
        return lead_import
    mapping = lead_import.column_mapping
    if not mapping.get("company_name"):
        raise ValidationFailedError("A Company Name column must be mapped before import.")

    organization_id = organization_id_of(lead_import)
    rows = raw_rows_of(db, lead_import)

    # Cross-campaign dedup index for the organization's OTHER campaigns (built once).
    dedup_index = build_dedup_index(db, organization_id, {lead_import.id})

    companies_by_key: dict[str, Company] = {}
    created, dropped, duplicates = _materialize_rows(
        db, lead_import, rows, mapping, dedup_index, companies_by_key
    )

    lead_import.status = ImportStatus.IMPORTED
    _purge_raw_rows(db, lead_import)  # drop the staged payload once materialized
    active_icp = next((i for i in lead_import.campaign.icps if i.is_active), None)
    if active_icp:
        lead_import.icp_snapshot_hash = icp_fingerprint(active_icp)
    db.commit()
    db.refresh(lead_import)
    logger.info(
        "Import %s confirmed: %s companies, %s leads, %s dropped, %s duplicates eliminated",
        lead_import.id, len(companies_by_key), created, dropped, duplicates,
    )
    return lead_import


def confirm_reupload(db: Session, pending: LeadImport, permanent: LeadImport) -> tuple[LeadImport, str]:
    """Confirm a pending re-upload into the campaign's permanent import.
    Returns (permanent, mode). Mode is re-resolved here — never trusted from the client."""
    from app.services.csv.reupload_service import (
        CampaignIdentityIndex,
        icp_fingerprint,
        reset_computed_results,
        resolve_upload_mode,
    )

    if pending.status != ImportStatus.MAPPING:
        raise ValidationFailedError("This upload was already processed.")
    if permanent.status in (ImportStatus.IMPORTED, ImportStatus.QUALIFYING, ImportStatus.SCORING):
        raise ValidationFailedError("A run is already in progress for this campaign — wait for it to finish.")
    mapping = pending.column_mapping
    if not mapping.get("company_name"):
        raise ValidationFailedError("A Company Name column must be mapped before import.")

    icp = next((i for i in permanent.campaign.icps if i.is_active), None)
    pending_rows = raw_rows_of(db, pending)
    resolution = resolve_upload_mode(db, permanent, icp, pending_rows, mapping)
    mode = resolution["mode"]
    if mode == "blocked":
        raise ValidationFailedError(
            "This file contains no new leads and your targeting inputs haven't changed — there is nothing to run."
        )

    if mode == "rerun":
        # Entities are kept; every computed artifact (qualification scores, lead scores,
        # drafts, manual decisions) is erased and the whole set re-runs against new inputs.
        reset_computed_results(db, permanent)

    # Row-level identity skip + campaign-aware dedup (never dedup against itself).
    organization_id = organization_id_of(permanent)
    identity = CampaignIdentityIndex.build(db, permanent.id)
    dedup_index = build_dedup_index(db, organization_id, {pending.id, permanent.id})

    # Seed with the run's existing companies so new leads at known companies attach to
    # the same canonical Company row.
    existing = db.scalars(
        select(Company)
        .join(CompanyQualification, CompanyQualification.company_id == Company.id)
        .where(CompanyQualification.lead_import_id == permanent.id)
    ).all()
    companies_by_key = {_company_key(c.name, c.domain): c for c in existing}

    created, dropped, _ = _materialize_rows(
        db, permanent, pending_rows, mapping, dedup_index, companies_by_key, skip_identity=identity
    )

    if icp:
        permanent.icp_snapshot_hash = icp_fingerprint(icp)
    permanent.status = ImportStatus.IMPORTED
    permanent.total_rows = permanent.total_rows + created
    db.delete(pending)
    db.commit()
    db.refresh(permanent)
    logger.info(
        "Re-upload confirmed into import %s (mode=%s): %s new leads, %s dropped",
        permanent.id, mode, created, dropped,
    )
    return permanent, mode
